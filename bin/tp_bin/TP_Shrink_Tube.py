import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from bin.tp_bin.model.TP_Tube_Model import TubeModelClass
from bin.tp_bin.model.TP.ShrinkTubeMandrel import DC0128_ADBT
from bin.tp_bin.model.TP.FormingMandrel import DC0128_ADBT_
from bin.tp_bin.model.TP.DrawMold import DC0128_ADIE
from bin.tp_bin.model.TP.FormingMold import DC0128_ADIE_
from bin.tp_bin.model.TP.UnloadingMold import DC0128_AD04
from bin.tp_bin.model.TP.CuttingClipMold1 import DC0128_AD01
from bin.tp_bin.model.TP.CuttingClipMold2 import DC0128_AD01_
from bin.tp_bin.utils.TP_SQLite_Control import MoldControl
import shutil
import zipfile
import subprocess
import os
import csv


# 壁厚特殊情况和是否打taper 会导致模具的选择有所不同
TP_MOLDS = {
        'apply_taper_true': ['裁剪夹模1', '裁剪夹模2', '抽管退料模', '抽管芯轴', '抽套模'],
        'apply_taper_false_special_thinkness_true': ['裁剪夹模1', '裁剪夹模2', '抽管退料模', '抽管芯轴', '抽套模', '成型抽套模', '成型抽管芯轴'],#形线加1mm
        'apply_taper_false_special_thinkness_false': ['裁剪夹模1', '裁剪夹模2', '抽管退料模', '抽管芯轴', '抽套模']
    }

class TPShrinkTubeClass:
    """TP缩管计算类
    这个类实现TP抽的所有功能，这是一个完整的TP抽管计算类。
    """

    def get_tube_params(self) -> dict:
        """将管件参数转换为字典，并返回"""
        params_dict = {}
        for i in range(len(self.tube_df_params)):
            for col_param, col_value in [('参数', '值')]:
                try:
                    if col_param in self.tube_df_params.columns and col_value in self.tube_df_params.columns:
                        params_dict[self.tube_df_params[col_param][i]] = self.tube_df_params[col_value][i]
                        break
                except Exception as e:
                    print(f"WARNING: 转换参数时出错: {str(e)}", " function: get_tube_params")
        return params_dict
    
    def get_tube_params_df(self):
        """增加参数描述"""
        parameter_list = list(self.tube_df_params['参数'])
        value_list = []
        for v in self.tube_df_params['值']:
            value_list.append(str(v))
        description_list = []
        for param in parameter_list:
            description_list.append(self.Parameter_Description['管件参数'][param])
        tube_params_df = pd.DataFrame({
            '参数': parameter_list,
            '值': value_list,
            '描述': description_list
        })
        return tube_params_df

    def __init__(self, file_path):
        """初始化TPShrinkTubeClass"""
        # 初始化Mold_Object
        self.Mold_Object = {}
        # 初始化tube_model为None
        self.tube_model = None

        # 导入模具的参数描述和计算方法
        with open('bin/tp_bin/model/TP/Parameter_Description.json', 'r', encoding='utf-8') as file:
            self.Parameter_Description = json.load(file)
        with open('bin/tp_bin/model/TP/Parameter_Calculate_Method.json', 'r', encoding='utf-8') as file:
            self.Parameter_Calculate_Method = json.load(file)

        if file_path:
            self.load_tube(file_path)
        else:
            print('未传入管件参数，无法计算，请使用load_tube()函数加载文件')

        self.tube_df_params = None
        print('TPShrinkTubeClass初始化完成。', 'function: __init__')

    def load_tube(self, file_path: str):
        """加载管件参数，更改文件路径"""
        self.tube_model = TubeModelClass(file_path)
        self.tube_df_params = self.tube_model.get_params()
        print(f"成功加载管件参数，参数表形状: {self.tube_df_params.shape}。", "function: load_tube")


    def calculate(self, user_name: str, apply_taper: bool, mold_list: list, special_thinkness: bool, interference_length: float):
        """计算模具的参数"""
        self.Mold_Object = {}

        # 准备外部参数
        # 获取管件参数，字典格式
        tube_params = self.get_tube_params()
        if not tube_params:
            print("错误: 无法获取管件参数")
            return None

        external_params = {
            '车种规格': tube_params.get('车种规格'),
            '设计者': user_name,
            '图号': ''  # 将在后面设置
        }

        # 创建模具对象并设置参数
        for mold_name in mold_list:
            # 获取图号
            graph_number = self.get_numbers(mold_name)
            external_params['图号'] = graph_number

            # 创建模具对象
            if mold_name == '裁剪夹模1':
                self.Mold_Object[mold_name] = DC0128_AD01()
            elif mold_name == '裁剪夹模2':
                self.Mold_Object[mold_name] = DC0128_AD01_()
            elif mold_name == '抽套模':
                self.Mold_Object[mold_name] = DC0128_ADIE()
            elif mold_name == '成型抽套模':
                self.Mold_Object[mold_name] = DC0128_ADIE_()
            elif mold_name == '抽管芯轴':
                self.Mold_Object[mold_name] = DC0128_ADBT(interference_length)
            elif mold_name == '成型抽管芯轴':
                self.Mold_Object[mold_name] = DC0128_ADBT_()
            elif mold_name == '抽管退料模':
                self.Mold_Object[mold_name] = DC0128_AD04()

            # 检查创建的对象
            mold = self.Mold_Object[mold_name]
            mold.set_params(self.tube_df_params, external_params)

        print(f"成功创建了 {len(self.Mold_Object)} 个模具对象。", "function: calculate")
        return self.Mold_Object

    def save_all(self):
        """将所有参数保存到数据库中"""
        sqlite_control = MoldControl()
        sqlite_control.insert_data(self.get_all_params())
        print('所有参数已保存到数据库中。', 'function: save_all')

    def get_numbers(self, mold_name):
        """通过查询数据库，获取图号"""
        # 定义模具名称与图号的映射
        self.Drawing_Number = {
            '裁剪夹模1': 'AD01',
            '裁剪夹模2': 'AD01_',
            '抽套模': 'ADIE',
            '成型抽套模': 'ADIE_',
            '抽管芯轴': 'ADBT',
            '成型抽管芯轴': 'ADBT_',
            '抽管退料模': 'AD04',
        }

        try:
            # big_graph_number
            sqlite_control = MoldControl()
            number = sqlite_control.query_max_graph_number('DC0128',self.Drawing_Number[mold_name])
            # print('查询到的图号为：', number)
            # 如果不存在记录，则会返回 0
            number = number + 1
            str_number = str(number)

            if len(str_number) != 4:
                new_number = self.Drawing_Number[mold_name] + '0' * (4 - len(str_number)) + str_number
            else:
                new_number = self.Drawing_Number[mold_name] + str_number
            return new_number
        except:
            print('没有找到{}的图号，将使用默认的图号'.format(mold_name))
            new_number = self.Drawing_Number[mold_name] + '0000'
            return new_number

    def modify_parameters(self, mold_name: str, key: str, value: str):
        """修改模具的参数"""
        try:
            mold = self.Mold_Object[mold_name]
            mold._modify_parameters(key, value)
        except Exception as e:
            temp_str = 'Error: 修改参数失败，位置位于：ShrinkTubeClass.modify_parameters'
            print(temp_str)

    def df_to_dict(self, df: pd.DataFrame) -> dict:
        """将DataFrame转换为字典"""
        Dict = {}
        for i in range(len(df)):
            Dict[df['参数'][i]] = df['值'][i]
        return Dict

    def get_all_params(self) -> dict:
        """获取当前计算的所有参数，返回嵌套字典"""
        # 将关键参数同时放入该字典中
        ALL_params = {
            '管件参数': self.df_to_dict(self.tube_df_params),
        }
        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            # 使用计算之后的图号作为关键key
            graph_number = self.df_to_dict(mold_params)['图号']
            ALL_params[graph_number] = self.df_to_dict(mold_params)

        return ALL_params

    def get_diameter(self,Parameter,image_number, param):
        try:
            return Parameter[image_number][param]
        except KeyError:
            param = '%%C' + param
            return Parameter[image_number][param]

    def get_molds_params_df(self):
        """获取当前计算的所有模具的参数，返回 df
        用于网页展示所有的模具数据。
        没有其余用法"""
        # 将关键参数同时放入该字典中
        ALL_params = {}

        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            image_number = mold_params.loc[mold_params['参数'] == '图号', '值'].values[0][:-4]

            mold_name = self.df_to_dict(mold_params)['模具名称']

            parameter_list = list(mold_params['参数'])
            value_list = []
            for v in mold_params['值']:
                value_list.append(str(v))
            description_list = []
            calclater_list = []
            for param in parameter_list:
                description_list.append(self.Parameter_Description[image_number][param])
                calclater_list.append(self.Parameter_Calculate_Method[image_number][param])

            params_df = pd.DataFrame({
                '参数': parameter_list,
                '值': value_list,
                '描述': description_list,
                '计算方法': calclater_list
            })
            ALL_params[mold_name] = params_df

        return ALL_params

    def output_dxf(self, output_path: str):
        """将所有模具的DXF文件保存到指定路径"""
        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold.modify_dxf() # 修改DXF文件
            mold_params = mold.get_params()
            mold_params_dict = self.df_to_dict(mold_params)
            mold.save_dxf(output_path, mold_params_dict['图号'])

    # 刘宇：增加 缩管模的step文件生成
    def output_gen_3d(self, output_path: str, gen3d_points):
        """将缩管模的step文件保存到指定路径"""
        # 确保目录存在
        if not os.path.exists(output_path):
            os.makedirs(output_path)
            print(f"创建目录: {output_path}")

        # 清理并删除目录下已存在的step文件，防止之前的文件干扰
        print(f"清理目录 {output_path} 中的旧step文件")
        for file in os.listdir(output_path):
            if file.endswith(".step"):
                file_path = os.path.join(output_path, file)
                try:
                    os.remove(file_path)
                    print(f"删除旧文件: {file_path}")
                except Exception as e:
                    print(f"删除文件时出错: {str(e)}")

        if gen3d_points.empty:
            print("没有生成的点，无法生成step文件")
            return

        gen3d_points = gen3d_points.values.tolist()

        point0 = gen3d_points[0]
        pointn = gen3d_points[-1]

        # points 应该是以(T_L+9,T_D/2)开始，以结束(9,T_d/2)的
        if point0[1] > pointn[1] and point0[0] > pointn[0]:
            T_D = point0[1] * 2
            T_d = pointn[1] * 2
            T_L = int(abs(point0[0]-pointn[0])) + 1
            gen3d_points = gen3d_points[::-1][:-1]
            gen3d_points.append([T_L + 9, T_D/2])
            points = gen3d_points[::-1]
        else:
            print("不满足基本条件，无法生成step文件")
            return

        # print(points)
        # 保存为 CSV 文件
        points_path = os.path.join(output_path, 'line_points.csv')
        with open(points_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            # 写入表头
            writer.writerow(['X', 'Y'])
            # 写入数据
            for point in points:
                writer.writerow(point)

        if os.path.exists('..\\make3D\\FreeCAD_0.21\\bin\\python.exe'):
            print('存在生成程序，将进行缩管模的生成')
            models_path = '..\\make3D\\models\\TaperShrinkTubeMold.py'
            output_path = os.path.join(output_path, 'TaperShrinkTubeMold.step')
            cmd = '\"..\\make3D\\FreeCAD_0.21\\bin\\python.exe\" ' + models_path + ' ' + output_path + ' ' + str(T_D) + ' ' + str(T_d) + ' ' + str(T_L) + ' ' + points_path
            # print(cmd)
            # Run the command
            process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            stdout, stderr = process.communicate()
            print('stdout:', stdout, 'stderr:', stderr)
            if stdout != '':
                print(stdout)
                return stdout
            elif stderr != '':
                print(stderr)
                return stderr
            print('成功生成step文件')
        else:
            print('不存在生成程序，将跳过缩管模的生成')
    # 刘宇 结束

    def output_excel(self, output_path: str):
        """将所有模具的参数保存到Excel文件中"""
        # 已经存在该文件，则删除
        excel_filename = os.path.join(output_path, '模具参数汇总.xlsx')
        if os.path.exists(excel_filename):
            os.remove(excel_filename)

        # 检查文件是否存在，如果不存在则创建
        if not os.path.exists(excel_filename):
            with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
                self.tube_df_params.to_excel(writer, sheet_name='管件参数', index=False)
                for mold_name in self.Mold_Object:
                    mold = self.Mold_Object[mold_name]
                    mold_params = mold.get_params()
                    mold_params.to_excel(writer, sheet_name=mold_name, index=False)

        # 加载Excel文件
        wb = load_workbook(excel_filename)

        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheet_name]
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            # 遍历当前工作表的所有单元格，设置居中对齐
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(excel_filename)
        wb.close()

    def output_zip_from_cache(self,out_root, gen3d_points):
        """将所有模具的DXF文件保存到指定路径"""
        if not os.path.exists(out_root):
            os.mkdir(out_root)
        output_path = str(os.path.join(out_root, self.get_tube_params()['车种规格']))
        zip_filename = str(os.path.join(out_root, self.get_tube_params()['车种规格'] + '.zip'))

        if not os.path.exists(output_path):
            os.mkdir(output_path)
        else:
            shutil.rmtree(output_path)
            os.mkdir(output_path)

        if os.path.exists(zip_filename):
            os.remove(zip_filename)

        self.output_dxf(output_path)
        self.output_excel(output_path)

        self.output_gen_3d(output_path, gen3d_points)

        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(output_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # 计算在压缩文件中的相对路径
                    arcname = os.path.relpath(file_path, output_path)
                    zipf.write(file_path, arcname)
        print('已经保存到：', zip_filename)
        return zip_filename
