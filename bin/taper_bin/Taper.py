import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from bin.taper_bin.model.TaperModel import TaperModelClass
from bin.taper_bin.model.Taper.TaperMold import EC0120_B000,EC0121_B000
from bin.taper_bin.model.Taper.TaperPositionMold import C000
from bin.taper_bin.model.Taper.TaperClampMold import J000
from bin.taper_bin.model.Taper.TaperBackPosMold import E000
from bin.taper_bin.model.Taper.TaperInnerPosMold import I000
from bin.taper_bin.model.Taper.TaperPullRodMold import H000
from bin.taper_bin.utils.Taper_SQLite_control import MoldControl
import shutil
import zipfile


Taper_Mold = {
    "EC0120":["Taper模两瓣", "定位杆", "夹模","后定位模","内定位模","拉杆"],
    "EC0121":["Taper模四瓣", "定位杆", "夹模"]
}
class TaperClass:
    """Taper计算类
     这个类实现TP抽的所有功能，这是一个完整的TP抽管计算类。
    """

    def get_tube_params(self) -> dict:
        """将管件参数转换为字典，并返回"""
        params_dict = {}
        for i in range(len(self.tube_df_params)):
            for col_param, col_value in [('参数', '值')]:
                try:
                    if col_param in self.tube_df_params.columns and col_value in self.tube_df_params.columns:
                        params_dict[self.tube_df_params[col_param][i]] = self.tube_df_params[col_value][i]
                        break
                except Exception as e:
                    print(f"WARNING: 转换参数时出错: {str(e)}", " function: get_tube_params")
        return params_dict

    def get_tube_params_df(self):
        """增加参数描述"""
        parameter_list = list(self.tube_df_params['参数'])
        value_list = []
        for v in self.tube_df_params['值']:
            value_list.append(str(v))
        description_list = []
        for param in parameter_list:
            description_list.append(self.Parameter_Description['管件参数'][param])
        tube_params_df = pd.DataFrame({
            '参数': parameter_list,
            '值': value_list,
            '描述': description_list
        })
        return tube_params_df

    def __init__(self, file_path: str):
        """初始化TaperClass"""
        # 初始化Mold_Object
        self.Mold_Object = {}
        # 初始化tube_model为None
        self.tube_model = None
        # 导入模具的参数描述和计算方法
        with open('bin/taper_bin/model/Taper/Parameter_Description.json', 'r', encoding='utf-8') as file:
            self.Parameter_Description = json.load(file)
        #print(self.Parameter_Description)  #调试语句
        with open('bin/taper_bin/model/Taper/Parameter_Calculate_Method.json', 'r', encoding='utf-8') as file:
            self.Parameter_Calculate_Method = json.load(file)

        if file_path:
            self.load_tube(file_path)
        else:
            print('未传入管件参数，无法计算，请使用load_tube()函数加载文件')

        self.tube_df_params = None
        print('TaperClass初始化完成。', 'function: __init__')

    def load_tube(self, file_path: str):
        """加载管件参数，更改文件路径"""
        self.tube_model = TaperModelClass(file_path)
        self.tube_df_params = self.tube_model.get_params()
        print(f"成功加载管件参数，参数表形状: {self.tube_df_params.shape}。", "function: load_tube")

    def get_numbers(self, mold_name, machine_type):
        """通过查询数据库，获取不带车种的图号"""

        # 定义模具名称与图号的映射
        self.Drawing_Number = {
            'Taper模两瓣': 'B000',
            'Taper模四瓣': 'B000',
            '定位杆': 'C000',
            '夹模': 'J000',
            '后定位模':'E000',
            '内定位模':'I000',
            '拉杆':'H000',
        }

        try:
            # big_graph_number
            sqlite_control = MoldControl()
            number = sqlite_control.query_max_graph_number(machine_type,self.Drawing_Number[mold_name])
            # 如果不存在记录，则会返回 0
            number = number + 1
            str_number = str(number)

            if len(str_number) != 4:
                new_number = self.Drawing_Number[mold_name] + '0' * (4 - len(str_number)) + str_number
            else:
                new_number = self.Drawing_Number[mold_name] + str_number
            return machine_type + '-' + new_number
        except:
            print('没有找到{}的图号，将使用默认的图号'.format(mold_name))
            new_number = self.Drawing_Number[mold_name] + '0000'
            return machine_type + '-' + new_number

    def calculate(self, user_name: str,machine_type:str, mold_list: list,tp_length,line_data,) -> dict:
        """计算模具的参数"""
        # 确保 tube_df_params 存在且格式正确
        self.Mold_Object = {}
        # 准备外部参数
        # 获取管件参数，字典格式
        tube_params = self.get_tube_params()
        if not tube_params:
            print("错误: 无法获取管件参数")
            return None

        external_params = {
            '车种规格': tube_params.get('车种规格'),
            '设计者': user_name,
            '图号': ''  # 将在后面设置
        }

        for mold_name in mold_list:
            # 获取图号
            graph_number = self.get_numbers(mold_name,machine_type)
            external_params['图号'] = graph_number

            # 创建模具对象
            if machine_type == 'EC0120':
                if mold_name == 'Taper模两瓣':
                    self.Mold_Object[mold_name] = EC0120_B000(tp_length,line_data)
                elif mold_name == '定位杆':
                    self.Mold_Object[mold_name] = C000()
                elif mold_name == '夹模':
                    self.Mold_Object[mold_name] = J000(machine_type)
                elif mold_name == '后定位模':
                    self.Mold_Object[mold_name] = E000()
                elif mold_name == '内定位模':
                    self.Mold_Object[mold_name] = I000()
                elif mold_name == '拉杆':
                    self.Mold_Object[mold_name] = H000()
            else:
                if mold_name == 'Taper模四瓣':
                    self.Mold_Object[mold_name] = EC0121_B000(tp_length,line_data)
                elif mold_name == '定位杆':
                    self.Mold_Object[mold_name] = C000()
                elif mold_name == '夹模':
                    self.Mold_Object[mold_name] = J000(machine_type)


            # 检查创建的对象
            mold = self.Mold_Object[mold_name]
            mold.set_params(self.tube_df_params, external_params)

        print(f"成功创建了 {len(self.Mold_Object)} 个模具对象。", "function: calculate")
        return self.Mold_Object

        return self.Mold_Object


    def modify_parameters(self, mold_name: str, key: str, value: str):
        """修改模具的参数"""
        # print('mold_name:', mold_name)
        # print('key:', key)
        # print('value:', value)
        try:
            mold = self.Mold_Object[mold_name]
            mold._modify_parameters(key, value)
        except Exception as e:
            temp_str = 'Error: 修改参数失败，位置位于：TaperTubeClass.modify_parameters'
            print(temp_str)

    def df_to_dict(self, df: pd.DataFrame) -> dict:
        """将DataFrame转换为字典"""
        Dict = {}
        for i in range(len(df)):
            #print(df)#调试语句
            try:
                Dict[df['Parameter'][i]] = df['Value'][i]
            except:
                Dict[df['参数'][i]] = df['值'][i]
        return Dict

    def get_all_params(self) -> dict:
        """获取当前计算的所有参数，返回嵌套字典"""
        # 将关键参数同时放入该字典中
        ALL_params = {
            '管件参数': self.df_to_dict(self.tube_df_params),
        }
        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            # 使用计算之后的图号作为关键key
            graph_number = self.df_to_dict(mold_params)['图号']
            ALL_params[graph_number] = self.df_to_dict(mold_params)

            #print(ALL_params)#调试语句

        return ALL_params

    def get_diameter(self,Parameter,image_number, param):
        try:
            return Parameter[image_number][param]
        except KeyError:
            param = '%%' + param
            return Parameter[image_number][param]

    def get_molds_params_df(self):
        """获取当前计算的所有模具的参数，返回 df"""

        ALL_params = {}

        for mold_name in self.Mold_Object:
            mold = self.Mold_Object[mold_name]
            mold_params = mold.get_params()
            #print(mold_params) #调试语句

            image_number = mold_params.loc[mold_params['参数'] == '图号', '值'].values[0][:-4]
            #print(image_number)  #调试语句
            mold_name = self.df_to_dict(mold_params)['模具名称']

            parameter_list = list(mold_params['参数'])
            value_list = []
            for v in mold_params['值']:
                value_list.append(str(v))
            description_list = []
            calclater_list = []
            for param in parameter_list:
                description_list.append(self.Parameter_Description[image_number][param])
                calclater_list.append(self.Parameter_Calculate_Method[image_number][param])

            params_df = pd.DataFrame({
                '参数': parameter_list,
                '值': value_list,
                '描述': description_list,
                '计算方法': calclater_list
            })
            ALL_params[mold_name] = params_df

        return ALL_params




    def save_all(self):
        """将所有参数保存到数据库中"""
        self.get_all_params()
        # 使用最新的方式保存数据
        sqlite_control = MoldControl()
        sqlite_control.insert_data(self.get_all_params())

    def output_dxf(self, output_path: str):
            """将所有模具的DXF文件保存到指定路径"""

            for mold_name in self.Mold_Object:
                mold = self.Mold_Object[mold_name]
                mold.modify_taper_dxf()
                mold_params = mold.get_params()
                mold_params_dict = self.df_to_dict(mold_params)
                mold.save_dxf(output_path, mold_params_dict['图号'])
            print('所有模具的DXF文件保存成功')


    def output_excel(self, output_path: str):
        """将所有模具的参数保存到Excel文件中"""
        #已经存在该文件，则删除
        excel_filename = os.path.join(output_path, '模具参数汇总.xlsx')
        if os.path.exists(excel_filename):
            os.remove(excel_filename)

        # 检查文件是否存在，如果不存在则创建
        if not os.path.exists(excel_filename):
            with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
                self.tube_df_params.to_excel(writer, sheet_name='管件参数', index=False)
                for mold_name in self.Mold_Object:
                    mold = self.Mold_Object[mold_name]
                    mold_params = mold.get_params()
                    mold_params.to_excel(writer, sheet_name=mold_name, index=False)

        # 加载Excel文件
        wb = load_workbook(excel_filename)

        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            # 获取当前工作表
            ws = wb[sheet_name]
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            # 遍历当前工作表的所有单元格，设置居中对齐
            for row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(excel_filename)
        wb.close()


    def output_zip_from_cache(self,out_root):
        """将所有模具的DXF文件保存到指定路径"""
        if not os.path.exists(out_root):
            os.mkdir(out_root)
        output_path = os.path.join(out_root, self.get_tube_params()['车种规格'])
        zip_filename = os.path.join(out_root,self.get_tube_params()['车种规格']+'.zip')

        if not os.path.exists(output_path):
            os.mkdir(output_path)
        else:
            shutil.rmtree(output_path)
            os.mkdir(output_path)

        if os.path.exists(zip_filename):
            os.remove(zip_filename)

        self.output_dxf(output_path)
        self.output_excel(output_path)

        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(output_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    # 计算在压缩文件中的相对路径
                    arcname = os.path.relpath(file_path, output_path)
                    zipf.write(file_path, arcname)

        print('已经保存到：', zip_filename)
        return zip_filename

